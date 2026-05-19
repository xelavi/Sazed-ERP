"""Shared utilities to generate XLSX exports across the ERP."""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
HEADER_ALIGN = Alignment(vertical='center', horizontal='left', wrap_text=True)
CELL_ALIGN = Alignment(vertical='center', horizontal='left', wrap_text=False)
NUMBER_FORMAT_CURRENCY = '#,##0.00'
DATE_FORMAT = 'yyyy-mm-dd'


def _coerce(value):
    """Convert Python values into types openpyxl handles cleanly."""
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, bool):
        return 'Sí' if value else 'No'
    return value


def build_xlsx_response(filename, sheet_title, headers, rows):
    """
    Generate an XLSX HttpResponse from a list of rows.

    headers: list of column titles.
    rows:    iterable of row tuples/lists (same length as headers).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or 'Export')[:31]

    # Header row
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, raw in enumerate(row, start=1):
            value = _coerce(raw)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = CELL_ALIGN
            if isinstance(value, (datetime, date)):
                cell.number_format = DATE_FORMAT

    # Reasonable auto-width based on content
    for c_idx, label in enumerate(headers, start=1):
        max_len = len(str(label))
        for r_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 48)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = filename if filename.endswith('.xlsx') else f'{filename}.xlsx'
    response = HttpResponse(
        buffer.read(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response
